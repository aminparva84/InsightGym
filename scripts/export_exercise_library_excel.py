"""
Export exercise library JSON files (Type 1/2/3) to formatted Excel files.
Run from project root: python scripts/export_exercise_library_excel.py
Outputs (docs/):
  - حرکات دستگاهی.xlsx
  - حرکات بدون دستگاه.xlsx
  - ترکیبی(دستگاه و فانکشنال).xlsx
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))
DATA_DIR = os.path.join(PROJECT_ROOT, "frontend", "src", "data")
DOCS_DIR = os.path.join(PROJECT_ROOT, "docs")

TYPE_FILES = [
    ("exerciseLibraryType1.json", "حرکات دستگاهی.xlsx", "Type 1"),
    ("exerciseLibraryType2.json", "حرکات بدون دستگاه.xlsx", "Type 2"),
    ("exerciseLibraryType3.json", "ترکیبی(دستگاه و فانکشنال).xlsx", "Type 3"),
]

HEADERS = [
    "Exercise (EN/FA)",
    "Exercise (FA)",
    "Hardness (1-10)",
    "Target muscle (EN/FA)",
    "Level",
    "Execution tips",
    "Breathing",
    "Gender",
]


def _load_groups(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _apply_layout(ws):
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="EEF2F7")
    header_font = Font(bold=True)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
    ws.row_dimensions[1].height = 24
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 60
    alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = alignment

    col_widths = {
        1: 34,  # Exercise EN/FA
        2: 24,  # Exercise FA
        3: 16,  # Hardness
        4: 28,  # Target
        5: 14,  # Level
        6: 60,  # Tips
        7: 20,  # Breathing
        8: 16,  # Gender
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_workbook(groups, title, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(HEADERS)
    for group in groups:
        for item in group.get("items", []):
            level = item.get("level_fa", "")
            if "مبتدی" in level and "حرفه‌ای" in level:
                hardness = 5
            elif "متوسط" in level and "حرفه‌ای" in level:
                hardness = 7
            elif "مبتدی" in level and "متوسط" in level:
                hardness = 4
            elif "حرفه‌ای" in level:
                hardness = 9
            elif "متوسط" in level:
                hardness = 6
            elif "مبتدی" in level:
                hardness = 2
            else:
                hardness = ""
            ws.append([
                f"{item.get('name_en', '')} / {item.get('name_fa', '')}",
                item.get("name_fa", ""),
                hardness,
                f"{item.get('target_group_en', '')} / {item.get('target_group_fa', '')}",
                item.get("level_fa", ""),
                item.get("tips_fa", ""),
                item.get("breathing_fa", ""),
                item.get("gender_fa", ""),
            ])
    _apply_layout(ws)
    os.makedirs(DOCS_DIR, exist_ok=True)
    wb.save(output_path)


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    for json_name, excel_name, sheet_title in TYPE_FILES:
        json_path = os.path.join(DATA_DIR, json_name)
        excel_path = os.path.join(DOCS_DIR, excel_name)
        groups = _load_groups(json_path)
        _write_workbook(groups, sheet_title, excel_path)
        print("Excel saved:", excel_path)


if __name__ == "__main__":
    main()
